# imports
import pandas as pd
import xlsxwriter
import openpyxl
from openpyxl import load_workbook
import os
import sys
from datetime import datetime


class ExcelGenerator:

    def __init__(self, data_path, layout_path):
        """
        :param data_path: absolute path of the data excel file
        :param layout_path: absolute path of the layout excel file
        """
        self.data_path = data_path
        self.layout_path = layout_path

    def construct_dfs(self):
        # File Names and for naming
        exp_name = 'Picogreen'
        plate_num = '1'
        pop_num = '1'
        report_name = generate_title(exp_name, plate_num, pop_num)

        # alter this to change concentration calculation and graph generation
        DNA_concentrations = [200, 100, 50, 25, 12.5, 6.25, 3.125, 0]

        # generates the i# variables corresponding to the layout and its sample names
        layout_df = load_spreadsheet(self.layout_path)
        load_tags(layout_df)

        # construct plate dataframe
        plate_df = load_spreadsheet(self.data_path, sheet_name='Result sheet', coord=[56, 152])
        plate_df.columns = ['<>', 'value']
        plate_df.index = plate_df.index - 54

        # The assigned positions from TECAN reader to a plate
        layout_to_tecan24 = {
            i1: (58, 61), i2: (70, 73), i3: (82, 85), i4: (94, 97),
            i5: (106, 109), i6: (118, 121), i7: (130, 133), i8: (142, 145),
            i9: (61, 64), i10: (73, 76), i11: (85, 88), i12: (97, 100),
            i13: (109, 112), i14: (121, 124), i15: (133, 135), i16: (145, 148),
            i17: (64, 67), i18: (76, 79), i19: (88, 91), i20: (100, 103),
            i21: (112, 115), i22: (124, 127), i23: (136, 139), i24: (148, 151),
            i25: (67, 70), i26: (79, 82), i27: (91, 94), i28: (103, 105),
            i29: (115, 118), i30: (127, 130), i31: (139, 142), i32: (151, 154)
        }

        # determine averages
        averages = pd.DataFrame.from_dict(find_avg(plate_df, layout_to_tecan24), orient='index')
        averages.columns = ['value']

        return report_name, layout_df, plate_df, layout_to_tecan24, averages, DNA_concentrations

    def excel_manipulations(self, report_name, DNA_concentrations, averages, layout_df, plate_df):

        # initialization setup and formatting
        workbook = xlsxwriter.Workbook(f'{report_name}.xlsx')
        worksheet = workbook.add_worksheet('std_dna')
        bold = workbook.add_format({'bold': 1})

        # Input data in excel
        headings = ['Concentration', 'Average Value']
        data = [DNA_concentrations, averages.iloc[0:8, :]['value'].tolist()]
        worksheet.write_row('A1', headings, bold)
        worksheet.write_column('A2', data[0])
        worksheet.write_column('B2', data[1])

        # create new scatter chart and add/ configure series
        std_dna_linear_chart = workbook.add_chart({'type': 'scatter'})
        std_dna_linear_chart.add_series({
            'name': '=std_dna!$B$1',
            'categories': '=std_dna!$A$2:$A$9',
            'trendline': {'type': 'linear', 'display_equation': True, 'line': {'color': 'red'}},
            'values': '=std_dna!$B$2:$B$9'})

        # labeling, title, chart style. Then insert into worksheet
        std_dna_linear_chart.set_title({'name': 'Average Picogreen Value'})
        std_dna_linear_chart.set_x_axis({'name': 'Concentration (ng/ml)'})
        std_dna_linear_chart.set_y_axis({'name': 'Value'})
        std_dna_linear_chart.set_style(11)
        worksheet.insert_chart('D2', std_dna_linear_chart, {'x_offset': 25, 'y_offset': 10})

        # Record the slope and intercept needed to solve for X
        worksheet.write(11, 0, 'Slope')
        worksheet.write(11, 1, 'Intercept')
        worksheet.write_formula('A13', '=SLOPE(B2:B9,A2:A9)')
        worksheet.write_formula('B13', '=INTERCEPT(B2:B9,A2:A9)')

        workbook.close()

        # Create sheets containing information on layout used and recorded plate values & calculated averages
        transfer_data(layout_df, report_name, 'plate_layout')
        transfer_data(plate_df, report_name, 'Picogreen_Values')
        transfer_data(averages, report_name, 'Sample_Averages')

        # Calculate concentration from Standard DNA correlation Trendline
        # The trendline will be reversed to X = (Y-B/M) to solve for concentration for all the picogreen values
        calculate(report_name)


def transfer_data(data_frame, xlsx_name, sheet_title):
    """
    Transfers existing panda dataframe to an Excel workbook. Adds a new sheet with desired name.
    xlsx_name does not include file type
    """
    relative_path = os.path.join(sys.path[0], f'{xlsx_name}.xlsx')
    book = load_workbook(relative_path)
    with pd.ExcelWriter(relative_path, engine='openpyxl') as writer:
        writer.book = book
        data_frame.to_excel(writer, sheet_name=sheet_title)
        writer.save()
        writer.close()


def generate_title(exp_name, plate_num, pop_num):
    return datetime.now().strftime('%Y-%m-%d_%Hh_%Mm_%Ss') + f'_{exp_name}_Plate{plate_num}_Pop{pop_num}_report'


def load_tags(data_frame, read_row=(1, 4, 7, 10)):
    """
    Can change to what rows it reads (if you want to load more sample types, assuming 3 recording per sample)
    """
    s_tags = []
    for row_num in read_row:
        s_tags += data_frame[row_num].tolist()
    # TODO: impliment solution that does not use globals
    for j in range(len(s_tags)):
        globals()['i{0}'.format(j + 1)] = s_tags[j]


def calculate(report_name, sheet_title='Sample_Averages', calc_type='concentration', dilution_factor=500):
    """
    Calculate concentration in ng/ul after accounting for dilution factor.
    """
    wb = openpyxl.load_workbook(f'{report_name}.xlsx')
    selected_sheet = wb[sheet_title]
    if calc_type == 'concentration':
        # for Concentration in ng/ul including dilution factor
        selected_sheet['C1'] = 'C (ng/ul)'
        for cell_obj in selected_sheet['C2:C33']:
            cell_obj[0].value = f'=0.001*{dilution_factor}*(INDIRECT("RC[-1]",0)-std_dna!B13)/(std_dna!A13)'
        # generate max and min
        selected_sheet['E1'] = 'Max:'
        selected_sheet['E2'] = 'Min:'
        selected_sheet['F1'] = '=MAX(C2:C33)'
        selected_sheet['F2'] = '=MIN(C2:C33)'
    wb.save(f'{report_name}.xlsx')


def load_spreadsheet(file_directory, sheet_name='Sheet1', coord=[]):
    dataframe = pd.read_excel(file_directory, sheet_name=sheet_name)
    return dataframe.drop('Unnamed: 0', axis=1) if not coord else dataframe.iloc[coord[0]:coord[1], 0:2]


def find_avg(plate_df, assigned_pos, valid_reading=3):
    """
    Value selects via iloc the 3 specific rows the sample is placed in the TECAN reader.
    Summate then divide by the # of valid readings to obtain the avg/sample.
    """
    # TODO: impliment changes to calculation when one or more of the three readings are incorrect
    return {key: plate_df.iloc[value[0] - 58:value[1] - 58, 0:2]['value'].sum() / valid_reading
            for key, value in assigned_pos.items()}


def main():
    abs_data_path = os.path.abspath(fr'{sys.argv[1]}')
    abs_layout_path = os.path.abspath(fr'{sys.argv[2]}')

    run = ExcelGenerator(abs_data_path, abs_layout_path)
    report_name, layout_df, plate_df, layout_to_tecan24, averages, DNA_concentrations = run.construct_dfs()
    run.excel_manipulations(report_name, DNA_concentrations, averages, layout_df, plate_df)


if __name__ == '__main__':
    main()
